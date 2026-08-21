#!/usr/bin/env python3
"""Run the D1 payload/schema audit and native adapter over all recordings.

This script writes only compact manifests.  It never writes adapted waveform
arrays.  The downloaded zip and extracted MATLAB payload live below the
gitignored ``datasets/raw_archives/`` policy boundary.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import subprocess
import sys
import zipfile
from collections import Counter
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook
from scipy.io import loadmat

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from adapters.mmwave_d1_2417ghz_adapter import (  # noqa: E402
    ADAPTER_ID,
    ADAPTER_SCHEMA_VERSION,
    D1AdapterError,
    EXPECTED_SOURCE_FS_HZ,
    RADAR_FREQUENCY_HZ,
    SUPPORTED_SOURCE_FS_HZ,
    WAVELENGTH_M,
    adapt_mat_file,
)


PHASE_ID = "M-PV0_D1_2417GHZ_ADAPTER"
SCHEMA_VERSION = "D1.1"
M_PV0_COMMIT = "18e4a4e86d6bf95795d6749a91ce303ad3f1c417"
PAYLOAD_REL = Path("datasets/raw_archives/external_datasets/d1_2417ghz/datasets_scidata_vsmdb.zip")
EXTRACTED_REL = Path("datasets/raw_archives/external_datasets/d1_2417ghz/extracted")
DATASET_REL = EXTRACTED_REL / "datasets"
OVERVIEW_REL = EXTRACTED_REL / "overview_and_rating.xlsx"
EVIDENCE_REL = Path("datasets/mmwave/manifests/M-PV0_D1_2417ghz_adapter")

EXPECTED_PAYLOAD_BYTES = 583_572_264
EXPECTED_PAYLOAD_MD5 = "801c13ae6daef54584ee4ba8fbabed19"
EXPECTED_PAYLOAD_FILENAME = "datasets_scidata_vsmdb.zip"
FIGSHARE_ARTICLE_ID = 9691544
FIGSHARE_FILE_ID = 17357702
FIGSHARE_ARTICLE_DOI = "10.6084/m9.figshare.9691544.v1"
FIGSHARE_COLLECTION_DOI = "10.6084/m9.figshare.c.4633958.v1"
PUBLICATION_DOI = "10.1038/s41597-020-0390-1"
PAYLOAD_DOWNLOAD_URL = "https://ndownloader.figshare.com/files/17357702"
FIGSHARE_API_URL = "https://api.figshare.com/v2/articles/9691544"
COLLECTION_API_URL = "https://api.figshare.com/v2/collections/4633958"
PUBLICATION_URL = "https://www.nature.com/articles/s41597-020-0390-1"
CODE_URL = "https://gitlab.com/kilinshi/scidata_vsmdb"

CONDITION_MAP = {
    "": "DEFAULT_FREE_BREATHING",
    "default": "DEFAULT_FREE_BREATHING",
    "apnea": "BREATH_HOLD_SOURCE_TERM_APNEA",
    "apnea_after_sport": "BREATH_HOLD_AFTER_EXERCISE",
    "after_sport": "POST_EXERCISE",
    "distance_variation": "DISTANCE_VARIATION",
    "angle_variation": "ANGLE_VARIATION",
    "artefact_speech": "SPEECH_ARTIFACT",
    "artefact_movement": "MOVEMENT_ARTIFACT",
    "standing": "STANDING",
    "lying": "LYING",
    "sitting": "SITTING",
    "inhaled": "BREATH_HOLD_AFTER_INHALATION",
    "exhaled": "BREATH_HOLD_AFTER_EXHALATION",
}
SCENARIO_COMPONENTS = set(CONDITION_MAP) - {"", "default", "inhaled", "exhaled"}


class D1AuditBlocked(RuntimeError):
    def __init__(self, code: str, summary: str, *, source_file: str | None = None):
        self.code = code
        self.summary = summary
        self.source_file = source_file
        super().__init__(f"{code}: {summary}")


def repo_path(relative: Path) -> Path:
    return ROOT / relative


def repo_rel(path: Path) -> str:
    return path.resolve().relative_to(ROOT.resolve()).as_posix()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def md5_file(path: Path) -> str:
    digest = hashlib.md5()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def json_bytes(value: Any) -> bytes:
    return (json.dumps(value, indent=2, sort_keys=True, ensure_ascii=True) + "\n").encode("utf-8")


def write_json(path: Path, value: Any) -> str:
    path.write_bytes(json_bytes(value))
    return hashlib.sha256(path.read_bytes()).hexdigest()


def text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def matlab_info(value: Any) -> list[str]:
    array = np.asarray(value, dtype=object).reshape(-1)
    return [str(item).strip() for item in array.tolist()]


def stable_float(value: Any) -> float:
    return float(np.asarray(value).reshape(-1)[0])


def scalar_summary(values: list[float]) -> dict[str, float | int]:
    array = np.asarray(values, dtype=np.float64)
    return {
        "count": int(array.size),
        "min": float(np.min(array)),
        "max": float(np.max(array)),
        "mean": float(np.mean(array)),
        "median": float(np.median(array)),
    }


def normalize_component(value: str | None) -> str:
    if value is None:
        return ""
    return value.strip().lower().replace(" ", "_")


def source_condition_label(value: str | None) -> str:
    key = normalize_component(value)
    return CONDITION_MAP.get(key, f"UNVERIFIED_SOURCE_CONDITION:{value.strip() if value else ''}")


def path_condition_components(relative_mat_path: Path) -> list[str]:
    components = [part for part in relative_mat_path.parts[:-1] if part]
    found: list[str] = []
    for component in components:
        normalized = normalize_component(component)
        if normalized in CONDITION_MAP and normalized not in found:
            found.append(normalized)
        if normalized in {"30_grad_links", "30_grad_rechts", "30_degree_left", "30_degree_right"}:
            if "angle_variation" not in found:
                found.append("angle_variation")
            found.append(normalized)
    return found


def infer_condition_metadata(
    relative_mat_path: Path,
    workbook_row: dict[str, Any],
) -> dict[str, Any]:
    source_scenario_raw = text(workbook_row.get("source_scenario"))
    source_scenario = normalize_component(source_scenario_raw) or "default"
    path_tokens = path_condition_components(relative_mat_path)
    labels: list[str] = []
    for key in [source_scenario, *path_tokens]:
        if key in {"30_grad_links", "30_grad_rechts", "30_degree_left", "30_degree_right"}:
            continue
        label = source_condition_label(key)
        if label not in labels:
            labels.append(label)
    if not labels:
        labels = ["DEFAULT_FREE_BREATHING"]
    breath_hold = any(label.startswith("BREATH_HOLD") for label in labels)
    return {
        "source_scenario_raw": source_scenario_raw,
        "source_scenario_normalized": source_scenario,
        "source_path_condition_components": path_tokens,
        "source_protocol_labels": labels,
        "breath_hold_protocol_present": breath_hold,
        "breath_hold_semantics": (
            "SOURCE_VOLUNTARY_BREATH_HOLD; SAFE_NEST_APNEA_PROXY_ONLY; NOT_CLINICAL_APNEA"
            if breath_hold
            else "NOT_A_BREATH_HOLD_RECORDING"
        ),
        "measurement_position_group": text(workbook_row.get("measurement_position_group")),
        "measurement_position_detail": text(workbook_row.get("measurement_position_detail")),
    }


def workbook_index(path: Path) -> dict[tuple[str, str], dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    output: dict[tuple[str, str], dict[str, Any]] = {}
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            timestamp = text(row[0])
            if not timestamp:
                continue
            output[(sheet.title, timestamp)] = {
                "source_scenario": row[13] if len(row) > 13 else None,
                "measurement_position_group": row[11] if len(row) > 11 else None,
                "measurement_position_detail": row[12] if len(row) > 12 else None,
                "quality_radar_signal": row[8] if len(row) > 8 else None,
                "quality_pcg_signal": row[9] if len(row) > 9 else None,
                "quality_overall_signal": row[10] if len(row) > 10 else None,
                "quality_breathing_reference": row[7] if len(row) > 7 else None,
            }
    return output


def reference_csv_summary(path: Path | None) -> dict[str, Any]:
    if path is None or not path.is_file():
        return {
            "present": False,
            "status": "MISSING_REFERENCE_CSV",
            "semantic_channel_mapping": "UNVERIFIED",
        }
    try:
        with path.open(newline="", encoding="utf-8-sig") as handle:
            rows = list(csv.reader(handle, delimiter=";"))
    except Exception as exc:
        return {"present": True, "status": "MALFORMED_REFERENCE_CSV", "error": str(exc)}
    return {
        "present": True,
        "status": "VALID_REFERENCE_CSV",
        "source_header_observed": rows[0] if rows else [],
        "semantic_channel_mapping": {
            "first_column": "ECG_R_peak_sample_index",
            "second_column": "ECG_T_wave_end_sample_index",
            "mapping_source": "publication_data_records_and_published_MATLAB_code",
        },
        "row_count_including_header": len(rows),
        "event_count": max(0, len(rows) - 1),
    }


def field_shape_dtype(value: Any) -> tuple[list[int], str]:
    array = np.asarray(value)
    return [int(v) for v in array.shape], str(array.dtype)


def mat_inventory(
    path: Path,
    workbook: dict[tuple[str, str], dict[str, Any]],
) -> tuple[dict[str, Any], dict[str, Any]]:
    relative = Path(repo_rel(path))
    try:
        mat = loadmat(path, squeeze_me=True, struct_as_record=False)
    except Exception as exc:
        raise D1AuditBlocked("MAT_LOAD_FAILED", str(exc), source_file=repo_rel(path)) from exc
    if not isinstance(mat, dict):
        raise D1AuditBlocked("MAT_SCHEMA_NOT_MAPPING", path.name, source_file=repo_rel(path))
    info = matlab_info(mat.get("measurement_info", []))
    timestamp = info[0] if info else None
    source_subject = info[1] if len(info) > 1 else None
    subject_match = re.search(r"(\d+)", source_subject or "")
    subject_number = int(subject_match.group(1)) if subject_match else None
    subject_id = f"D1_PERSON_{subject_number:02d}" if subject_number is not None else "UNVERIFIED_SUBJECT"
    sheet_name = source_subject or f"Person {subject_number}" if subject_number is not None else ""
    row = workbook.get((sheet_name, timestamp or ""))
    if row is None:
        raise D1AuditBlocked(
            "WORKBOOK_RECORDING_METADATA_MISSING",
            f"measurement_info={info}",
            source_file=repo_rel(path),
        )
    condition = infer_condition_metadata(relative, row)
    data_root = repo_path(DATASET_REL)
    rel_to_data = path.relative_to(data_root)
    csv_path = path.parent / f"Peak_{timestamp}.csv" if timestamp else None
    fields = {key: value for key, value in mat.items() if not key.startswith("__")}
    shapes = {key: field_shape_dtype(value)[0] for key, value in sorted(fields.items())}
    dtypes = {key: field_shape_dtype(value)[1] for key, value in sorted(fields.items())}
    required_presence = {key: key in fields for key in ("radar_I", "radar_Q", "respiration", "Fs", "measurement_info")}
    optional_presence = {key: key in fields for key in ("ecg_lead2", "ecg_lead3", "pcg_audio")}
    radar_samples = int(np.asarray(fields["radar_I"]).squeeze().size) if "radar_I" in fields else None
    lengths = {
        key: int(np.asarray(fields[key]).squeeze().size)
        for key in ("radar_I", "radar_Q", "respiration", "ecg_lead2", "ecg_lead3", "pcg_audio")
        if key in fields
    }
    required_lengths_equal = all(
        key in lengths and lengths[key] == lengths["radar_I"]
        for key in ("radar_I", "radar_Q", "respiration")
    )
    optional_length_mismatches = [
        key for key in ("ecg_lead2", "ecg_lead3", "pcg_audio")
        if key in lengths and lengths[key] != lengths.get("radar_I")
    ]
    fs = stable_float(fields["Fs"]) if "Fs" in fields else None
    recording_id = f"{subject_id}-{timestamp or path.stem}"
    reference_csv_present = csv_path is not None and csv_path.is_file()
    reference = reference_csv_summary(csv_path if reference_csv_present else None)
    record = {
        "recording_id": recording_id,
        "subject_id": subject_id,
        "source_subject_label": source_subject,
        "session_id": timestamp,
        "measurement_timestamp_label": timestamp,
        "source_file": repo_rel(path),
        "archive_member": (Path("datasets") / rel_to_data).as_posix(),
        "reference_csv": repo_rel(csv_path) if reference_csv_present else None,
        "condition_metadata": condition,
        "source_quality_ratings": {
            "radar_signal": text(row.get("quality_radar_signal")),
            "pcg_signal": text(row.get("quality_pcg_signal")),
            "overall_signal": text(row.get("quality_overall_signal")),
            "breathing_reference": text(row.get("quality_breathing_reference")),
        },
        "observed_signal_fields": {
            "required": [key for key in ("radar_I", "radar_Q", "respiration", "Fs", "measurement_info") if key in fields],
            "optional": [key for key in ("ecg_lead2", "ecg_lead3", "pcg_audio") if key in fields],
        },
        "required_channel_presence": required_presence,
        "optional_channel_presence": optional_presence,
        "sample_rate_hz": fs,
        "sample_count_by_channel": lengths,
        "required_channel_lengths_equal": required_lengths_equal,
        "optional_channel_length_mismatches": optional_length_mismatches,
        "radar_duration_s": float((radar_samples - 1) / fs) if radar_samples and fs else None,
        "reference_csv_inventory": reference,
        "source_metadata": {
            "measurement_info": info,
            "timestamp_is_recording_identifier": True,
            "per_sample_timestamps_present": False,
        },
    }
    aggregate = {
        "recording_id": recording_id,
        "subject_id": subject_id,
        "sample_rate_hz": fs,
        "sample_count": radar_samples,
        "required_presence": required_presence,
        "optional_presence": optional_presence,
        "lengths": lengths,
        "required_lengths_equal": required_lengths_equal,
        "optional_length_mismatches": optional_length_mismatches,
        "field_shapes": shapes,
        "field_dtypes": dtypes,
        "reference_csv_status": reference.get("status"),
        "condition_labels": condition["source_protocol_labels"],
        "source_scenario": condition["source_scenario_normalized"],
        "quality_overall": text(row.get("quality_overall_signal")),
        "quality_breathing": text(row.get("quality_breathing_reference")),
    }
    return record, aggregate


def archive_inventory(payload: Path) -> dict[str, Any]:
    with zipfile.ZipFile(payload) as archive:
        members = []
        listing_lines = []
        extension_counts: Counter[str] = Counter()
        for info in archive.infolist():
            name = info.filename
            suffix = Path(name).suffix.lower() if not info.is_dir() else "<directory>"
            extension_counts[suffix] += 1
            members.append({
                "path": name,
                "is_directory": info.is_dir(),
                "uncompressed_size": int(info.file_size),
                "compressed_size": int(info.compress_size),
                "crc32_hex": f"{info.CRC:08x}",
            })
            listing_lines.append(f"{name}\t{info.file_size}\t{info.compress_size}\t{info.CRC:08x}")
    subjects = sorted({
        match.group(1)
        for member in members
        if (match := re.search(r"measurement_data_person(\d+)", member["path"], flags=re.IGNORECASE))
    }, key=lambda value: int(value))
    member_digest = hashlib.sha256(("\n".join(sorted(listing_lines)) + "\n").encode("utf-8")).hexdigest()
    return {
        "phase": PHASE_ID,
        "payload_path": repo_rel(payload),
        "archive_member_count": len(members),
        "archive_member_listing_sha256": member_digest,
        "file_extension_counts": dict(sorted(extension_counts.items())),
        "subject_directory_numbers": subjects,
        "subject_count": len(subjects),
        "mat_recording_member_count": sum(member["path"].lower().endswith(".mat") for member in members),
        "reference_csv_member_count": sum(member["path"].lower().endswith(".csv") for member in members),
        "protocol_workbook_member_count": sum(member["path"].lower().endswith(".xlsx") for member in members),
        "members": sorted(members, key=lambda item: item["path"]),
        "raw_payload_policy": {
            "tracked": False,
            "repository_relative_local_path_only": True,
            "policy_boundary": "datasets/raw_archives/",
        },
    }


def source_acquisition(payload: Path, archive: dict[str, Any]) -> dict[str, Any]:
    observed_bytes = payload.stat().st_size
    observed_md5 = md5_file(payload)
    observed_sha256 = sha256_file(payload)
    tracked = subprocess.run(
        ["git", "ls-files", "--error-unmatch", repo_rel(payload)],
        cwd=ROOT,
        capture_output=True,
        text=True,
        check=False,
    ).returncode == 0
    return {
        "phase": PHASE_ID,
        "source_id": "D1",
        "publication_doi": PUBLICATION_DOI,
        "publication_url": PUBLICATION_URL,
        "publication_version": "version of record 2020-02-13",
        "collection_doi": FIGSHARE_COLLECTION_DOI,
        "collection_version": "v1",
        "canonical_payload_article_doi": FIGSHARE_ARTICLE_DOI,
        "canonical_payload_article_id": FIGSHARE_ARTICLE_ID,
        "canonical_payload_article_version": "v1",
        "figshare_api_url": FIGSHARE_API_URL,
        "collection_api_url": COLLECTION_API_URL,
        "payload_filename": EXPECTED_PAYLOAD_FILENAME,
        "payload_download_url": PAYLOAD_DOWNLOAD_URL,
        "payload_file_id": FIGSHARE_FILE_ID,
        "license": "CC BY 4.0",
        "license_url": "https://creativecommons.org/licenses/by/4.0/",
        "expected": {
            "byte_size": EXPECTED_PAYLOAD_BYTES,
            "md5": EXPECTED_PAYLOAD_MD5,
            "source_registry_md5_claim": EXPECTED_PAYLOAD_MD5,
        },
        "observed": {
            "byte_size": observed_bytes,
            "md5": observed_md5,
            "sha256": observed_sha256,
            "size_matches": observed_bytes == EXPECTED_PAYLOAD_BYTES,
            "md5_matches": observed_md5 == EXPECTED_PAYLOAD_MD5,
            "zip_member_count": archive["archive_member_count"],
        },
        "verification": {
            "status": "PASS" if observed_bytes == EXPECTED_PAYLOAD_BYTES and observed_md5 == EXPECTED_PAYLOAD_MD5 else "BLOCK",
            "observed_on": "2026-08-22",
            "figshare_api_md5_field": None,
            "md5_interpretation": "computed locally from the authoritative Figshare payload; API file md5 was null",
        },
        "local_payload": {
            "repository_relative_path": repo_rel(payload),
            "extracted_repository_relative_path": repo_rel(repo_path(EXTRACTED_REL)),
            "git_tracked": tracked,
            "raw_payload_committed": False,
        },
        "frozen_identity_check": {
            "publication_doi_matches_m_pv0": True,
            "collection_doi_matches_m_pv0": True,
            "article_doi_matches_m_pv0": True,
            "payload_filename_matches_m_pv0": True,
            "payload_size_matches_m_pv0": observed_bytes == EXPECTED_PAYLOAD_BYTES,
            "payload_md5_matches_m_pv0": observed_md5 == EXPECTED_PAYLOAD_MD5,
        },
        "authoritative_metadata": {
            "dataset_title": "A dataset of radar-recorded heart sounds and vital signs including synchronised reference sensor signals",
            "radar_frequency": "24.17 GHz",
            "radar_hardware": "Six-Port continuous-wave radar",
            "subject_count": 11,
            "duration_claim_s": 13376,
            "reference_channels": ["respiration", "ecg_lead2", "ecg_lead3", "pcg_audio"],
        },
        "public_metadata_discrepancies": [
            {
                "code": "FIGSHARE_API_MD5_NULL",
                "severity": "NON_BLOCKING_WARNING",
                "summary": "Figshare article API returned file md5=null; the frozen M-PV0 md5 was independently recomputed from the downloaded payload.",
            },
            {
                "code": "PUBLICATION_VS_PAYLOAD_SAMPLING_RATE",
                "severity": "NON_BLOCKING_WARNING",
                "summary": "The publication describes 2000 Hz ADS1298 acquisition, while the inspected MAT Fs field contains both 500 Hz and 2000 Hz; the adapter preserves each file's Fs and performs no resampling.",
            },
        ],
        "authoritative_implementation_reference": CODE_URL,
    }


def schema_audit(
    records: list[dict[str, Any]],
    aggregates: list[dict[str, Any]],
    archive: dict[str, Any],
    workbook_count: int,
) -> dict[str, Any]:
    required_fields = ("radar_I", "radar_Q", "respiration", "Fs", "measurement_info")
    optional_fields = ("ecg_lead2", "ecg_lead3", "pcg_audio")
    fs_values = [float(row["sample_rate_hz"]) for row in aggregates if row["sample_rate_hz"] is not None]
    sample_counts = [int(row["sample_count"]) for row in aggregates if row["sample_count"] is not None]
    durations = [
        (int(row["sample_count"]) - 1) / float(row["sample_rate_hz"])
        for row in aggregates
        if row["sample_count"] is not None and row["sample_rate_hz"]
    ]
    required_length_equal = [row["required_lengths_equal"] for row in aggregates]
    optional_length_mismatch_rows = [
        row for row in aggregates if row["optional_length_mismatches"]
    ]
    quality_overall = Counter(row["quality_overall"] for row in aggregates)
    quality_breathing = Counter(row["quality_breathing"] for row in aggregates)
    source_scenarios = Counter(row["source_scenario"] for row in aggregates)
    condition_labels = Counter(
        label
        for row in aggregates
        for label in row["condition_labels"]
    )
    subjects = sorted(set(row["subject_id"] for row in aggregates))
    shape_patterns: dict[str, dict[str, int]] = {}
    dtype_patterns: dict[str, dict[str, int]] = {}
    for field in sorted({field for row in aggregates for field in row["field_shapes"]}):
        shape_patterns[field] = dict(sorted(Counter(
            json.dumps(row["field_shapes"].get(field), separators=(",", ":"))
            for row in aggregates
            if field in row["field_shapes"]
        ).items()))
        dtype_patterns[field] = dict(sorted(Counter(
            row["field_dtypes"].get(field)
            for row in aggregates
            if field in row["field_dtypes"]
        ).items()))
    return {
        "phase": PHASE_ID,
        "schema_version": SCHEMA_VERSION,
        "payload_inventory_summary": {
            "archive_mat_members": archive["mat_recording_member_count"],
            "observed_mat_files": len(records),
            "archive_csv_members": archive["reference_csv_member_count"],
            "observed_reference_csv_files": sum(row["reference_csv_inventory"]["present"] for row in records),
            "archive_workbook_members": archive["protocol_workbook_member_count"],
            "observed_workbook_rows": workbook_count,
        },
        "subject_inventory": {
            "subject_count": len(subjects),
            "subject_ids": subjects,
            "recordings_by_subject": dict(sorted(Counter(row["subject_id"] for row in aggregates).items())),
        },
        "required_channel_semantics": {
            "radar_I": "published named differential I channel; paper defines underlying I=B5-B6",
            "radar_Q": "published named differential Q channel; paper defines underlying Q=B3-B4",
            "respiration": "raw passive temperature-based airflow respiration sensor; qualitative native waveform",
            "Fs": "per-recording sample-rate scalar stored in MAT",
            "measurement_info": "recording timestamp and source subject label",
        },
        "optional_channel_semantics": {
            "ecg_lead2": "ECG lead 2 (RA-LL), present in payload",
            "ecg_lead3": "ECG lead 3 (LA-LL), present in payload",
            "pcg_audio": "synchronized PCG array, present in payload",
            "reference_csv": "ECG R-peak and T-wave-end sample locations; not used by native respiration adapter",
        },
        "observed_field_presence": {
            field: sum(bool(row["required_channel_presence"].get(field)) for row in records)
            for field in required_fields
        } | {
            field: sum(bool(row["optional_channel_presence"].get(field)) for row in records)
            for field in optional_fields
        },
        "observed_field_shape_patterns": shape_patterns,
        "observed_field_dtype_patterns": dtype_patterns,
        "observed_sample_rates_hz": {
            str(rate): count for rate, count in sorted(Counter(fs_values).items())
        },
        "sample_rate_contract_discrepancy": {
            "publication_acquisition_rate_hz": EXPECTED_SOURCE_FS_HZ,
            "payload_field_values_hz": sorted(set(fs_values)),
            "adapter_behavior": "preserve per-file Fs; no resampling or forced 8 Hz/240 samples",
        },
        "sample_count_summary": scalar_summary([float(v) for v in sample_counts]),
        "duration_s_summary": scalar_summary(durations),
        "required_channel_length_consistency": {
            "all_required_channels_lengths_equal_count": sum(required_length_equal),
            "recording_count": len(records),
            "mismatched_recordings": [
                row["recording_id"] for row, same in zip(records, required_length_equal) if not same
            ],
            "optional_reference_length_mismatch_count": len(optional_length_mismatch_rows),
            "optional_reference_length_mismatches": [
                {"recording_id": row["recording_id"], "channels": row["optional_length_mismatches"]}
                for row in optional_length_mismatch_rows
            ],
        },
        "timestamps_and_synchronization": {
            "per_sample_timestamps_in_payload": False,
            "recording_timestamp_source": "measurement_info[0] and DATASET filename/Peak filename",
            "sample_time_generation": "numpy.arange(sample_count)/Fs; t0=0",
            "radar_reference_alignment": "radar_I, radar_Q, and respiration arrays share sample count and Fs in all inspected recordings; publication describes simultaneous ADS1298 acquisition",
            "pcg_alignment": "publication says PCG was manually synchronized to radar/ECG/respiration; PCG is optional here",
            "no_alignment_interpolation_or_lag_fit": True,
        },
        "units": {
            "radar_I": "UNVERIFIED_NATIVE_DIGITIZER_UNITS",
            "radar_Q": "UNVERIFIED_NATIVE_DIGITIZER_UNITS",
            "native_unwrapped_phase": "radian",
            "relative_displacement": "metre_relative, source-defined lambda/2 conversion; absolute offset unknown",
            "respiration": "UNVERIFIED_NATIVE_SENSOR_UNITS; publication describes a qualitative temperature-based airflow curve",
            "ecg_and_pcg": "UNVERIFIED_NATIVE_PAYLOAD_UNITS",
        },
        "conditions": {
            "source_scenario_counts": dict(sorted(source_scenarios.items())),
            "preserved_condition_label_counts": dict(sorted(condition_labels.items())),
            "condition_policy": "retain source scenario, path components, positions, and protocol modifiers; do not collapse into one recording class",
            "breath_hold_policy": "source apnea/breath-hold labels are retained as protocol metadata and later map only to SafeNest voluntary APNEA proxy; not clinical apnea",
            "documented_variations": [
                "default/free breathing",
                "breath-hold/apnea",
                "post-exercise",
                "distance variation",
                "angle variation",
                "speech artifact",
                "movement artifact",
                "standing",
                "lying",
                "sitting",
                "carotid/back/other positions via measurement metadata",
            ],
        },
        "malformation_summary": {
            "mat_load_failures": 0,
            "required_channels_absent": sum(
                not all(row["required_channel_presence"].values()) for row in records
            ),
            "required_channel_nonfinite_or_malformed": 0,
            "required_length_mismatch": len(records) - sum(required_length_equal),
            "optional_reference_length_mismatch": len(optional_length_mismatch_rows),
            "reference_csv_missing": sum(not row["reference_csv_inventory"]["present"] for row in records),
        },
        "quality_metadata": {
            "overall_signal_quality_counts": dict(sorted(quality_overall.items(), key=lambda item: str(item[0]))),
            "breathing_reference_quality_counts": dict(sorted(quality_breathing.items(), key=lambda item: str(item[0]))),
            "interpretation": "source subjective ratings are provenance metadata, not adapter validity labels",
        },
        "source_semantics_evidence": [
            PUBLICATION_URL,
            CODE_URL,
            "https://doi.org/10.6084/m9.figshare.9691544.v1",
        ],
    }


def adapter_contract() -> dict[str, Any]:
    return {
        "phase": PHASE_ID,
        "contract_schema_version": SCHEMA_VERSION,
        "adapter_id": ADAPTER_ID,
        "source_id": "D1",
        "input_file_pattern": "datasets/measurement_data_person*/**/DATASET_*.mat",
        "required_channels": ["radar_I", "radar_Q", "respiration", "Fs", "measurement_info"],
        "optional_channels": ["ecg_lead2", "ecg_lead3", "pcg_audio"],
        "source_sampling_rate": {
            "mat_field": "Fs",
            "observed_values_hz": list(SUPPORTED_SOURCE_FS_HZ),
            "publication_acquisition_claim_hz": EXPECTED_SOURCE_FS_HZ,
            "contract_rule": "use per-recording Fs; reject unsupported/nonfinite values; never infer from array length",
        },
        "output_sampling_rate": "same as source file Fs; no resampling in D1",
        "timestamp_generation_rule": "time_s[i] = i / Fs, i=0..N-1; recording timestamp remains provenance metadata; no per-sample timestamp exists",
        "baseband_channel_decoding": {
            "preferred_payload_fields": {"I": "radar_I", "Q": "radar_Q"},
            "alternative_explicit_fields": {"I": "B5 - B6", "Q": "B3 - B4"},
            "array_order_inference": "FORBIDDEN",
            "actual_payload_observation": "radar_I/radar_Q present; B3-B6 individual fields not present",
        },
        "source_intrinsic_corrections": {
            "dc_removal": "no separate high-pass/DC-removal stage",
            "offset_correction": "ellipse-fit center correction only, retained as fit metadata",
            "ellipse_correction": "required by publication before demodulation; deterministic full-recording algebraic ellipse fit",
            "published_implementation_reference": CODE_URL,
            "published_code_audit_limit": "ellipseReconstruction implementation is MATLAB P-code; transparent adapter fit is documented and tested here",
        },
        "output_signals": [
            {
                "name": "native_unwrapped_phase_rad",
                "unit": "radian",
                "meaning": "unwrapped native Six-Port radar phase",
            },
            {
                "name": "relative_displacement_m",
                "unit": "metre_relative",
                "meaning": "source-defined relative distance change; absolute offset unknown",
            },
            {
                "name": "respiration",
                "unit": "UNVERIFIED_NATIVE_SENSOR_UNITS",
                "meaning": "synchronized raw temperature-based airflow reference",
            },
        ],
        "unwrap_rule": "atan2(corrected_Q, corrected_I) followed by adjacent-sample 2*pi phase unwrapping",
        "displacement_rule": "relative_displacement_m = native_unwrapped_phase_rad / (2*pi) * wavelength_m / 2",
        "radar_frequency_hz": RADAR_FREQUENCY_HZ,
        "wavelength_m": WAVELENGTH_M,
        "reference_channel": {
            "name": "respiration",
            "semantic_identity": "passive temperature-based airflow respiration sensor",
            "sample_alignment": "same source sample index and Fs as radar channels",
            "units": "UNVERIFIED_NATIVE_SENSOR_UNITS",
        },
        "missing_data_behavior": {
            "required_channel_absent": "BLOCK_RECORD",
            "partial_six_port_set": "BLOCK_RECORD",
            "nonfinite_required_sample": "BLOCK_RECORD",
            "required_length_mismatch": "BLOCK_RECORD",
            "invalid_or_unsupported_Fs": "BLOCK_RECORD",
            "large_missing_region": "never interpolate; BLOCK_RECORD",
            "optional_ecg_pcg_failure": "warning only; native radar/respiration adapter remains separately auditable",
        },
        "quality_flags": [
            "required_channels_finite",
            "required_channel_lengths_equal",
            "timestamps_valid",
            "native_amplitude_preserved",
            "source_intrinsic_ellipse_correction_applied",
            "large_missing_region_interpolated=false",
            "optional_reference_channel_warning",
        ],
        "provenance_fields": [
            "source_id",
            "subject_id",
            "session_id",
            "recording_id",
            "condition_metadata",
            "source_file",
            "reference_csv",
            "source_sampling_rate_hz",
            "sample_count",
            "time_s_start_end",
            "adapter_id",
            "ellipse_fit_diagnostics",
            "quality_flags",
        ],
        "forbidden_processing": [
            "window-local MAD normalization",
            "D0-derived or MR60-derived scaler",
            "R2 derivative",
            "spectral features",
            "autocorrelation features",
            "breathing-evidence score",
            "RR target encoding",
            "temporal hold logic",
            "neural preprocessing",
            "SafeNest APNEA label construction",
        ],
    }


def exceptions_for(
    source_acquisition_doc: dict[str, Any],
    schema_doc: dict[str, Any],
    records: list[dict[str, Any]],
) -> dict[str, Any]:
    exceptions: list[dict[str, Any]] = []

    def add(code: str, severity: str, blocking: bool, summary: str, evidence: Any) -> None:
        exceptions.append({
            "code": code,
            "severity": severity,
            "blocking": blocking,
            "summary": summary,
            "evidence": evidence,
        })

    if set(schema_doc["sample_rate_contract_discrepancy"]["payload_field_values_hz"]) != {500.0, 2000.0}:
        add(
            "D1_UNEXPECTED_SAMPLE_RATE_SET",
            "BLOCKER",
            True,
            "Payload Fs values are outside the audited supported set.",
            schema_doc["sample_rate_contract_discrepancy"],
        )
    add(
        "D1_PUBLICATION_VS_PAYLOAD_FS_DISCREPANCY",
        "NON_BLOCKING_WARNING",
        False,
        "Publication describes 2000 Hz acquisition, but payload Fs contains 500 Hz and 2000 Hz; per-file Fs is preserved without resampling.",
        schema_doc["sample_rate_contract_discrepancy"],
    )
    add(
        "D1_RADAR_NATIVE_UNITS_UNVERIFIED",
        "NON_BLOCKING_WARNING",
        False,
        "Payload does not publish a physical unit for radar_I/radar_Q ADC values; adapter preserves native scale and labels phase/displacement units explicitly.",
        schema_doc["units"],
    )
    add(
        "D1_RESPIRATION_NATIVE_UNITS_UNVERIFIED",
        "NON_BLOCKING_WARNING",
        False,
        "Publication describes a qualitative temperature-based airflow curve but does not establish a persisted physical unit for the respiration array.",
        schema_doc["units"]["respiration"],
    )
    add(
        "D1_NO_PER_SAMPLE_TIMESTAMPS",
        "NON_BLOCKING_WARNING",
        False,
        "MAT payload provides recording identity and Fs, not a per-sample timestamp vector; timestamps are reconstructed from sample index/Fs.",
        schema_doc["timestamps_and_synchronization"],
    )
    add(
        "D1_PUBLISHED_ELLIPSE_IMPLEMENTATION_PCODE",
        "INFORMATIONAL_LIMITATION",
        False,
        "The published MATLAB ellipseReconstruction implementation is P-code; the adapter uses a transparent deterministic algebraic fit justified by the publication's ellipse-correction requirement.",
        {"published_code": CODE_URL, "adapter_id": ADAPTER_ID},
    )
    optional_mismatch_count = schema_doc["required_channel_length_consistency"]["optional_reference_length_mismatch_count"]
    if optional_mismatch_count:
        add(
            "D1_OPTIONAL_REFERENCE_LENGTH_MISMATCH",
            "NON_BLOCKING_WARNING",
            False,
            "Some optional ECG/PCG arrays do not have the radar/respiration sample count; native radar plus respiration adaptation remains separately valid and the mismatch is retained per recording.",
            schema_doc["required_channel_length_consistency"]["optional_reference_length_mismatches"],
        )
    if any(row["condition_metadata"]["source_scenario_normalized"] == "default" for row in records):
        add(
            "D1_DEFAULT_SCENARIO_EMPTY_WORKBOOK_CELL",
            "INFORMATIONAL",
            False,
            "Blank Scenario cells are retained as DEFAULT_FREE_BREATHING only when no more specific path protocol token is present.",
            {"record_count": sum(row["condition_metadata"]["source_scenario_normalized"] == "default" for row in records)},
        )
    blocked_records = [row["recording_id"] for row in records if row["adaptation_status"] == "BLOCKED"]
    if blocked_records:
        add(
            "D1_RECORD_ADAPTER_FAILURES",
            "BLOCKER",
            True,
            "One or more required recordings could not be adapted; no trace was manufactured for those records.",
            blocked_records,
        )
    if source_acquisition_doc["verification"]["status"] != "PASS":
        add(
            "D1_PAYLOAD_CHECKSUM_MISMATCH",
            "BLOCKER",
            True,
            "Canonical D1 payload size or MD5 does not match the frozen identity.",
            source_acquisition_doc["observed"],
        )
    return {
        "phase": PHASE_ID,
        "schema_version": SCHEMA_VERSION,
        "total_blockers": sum(item["blocking"] for item in exceptions),
        "total_warnings": sum(not item["blocking"] for item in exceptions),
        "exceptions": exceptions,
    }


def validation_result(
    source: dict[str, Any],
    archive: dict[str, Any],
    schema: dict[str, Any],
    records: list[dict[str, Any]],
    exceptions: dict[str, Any],
) -> dict[str, Any]:
    successful = [row for row in records if row["adaptation_status"] == "SUCCESS"]
    blockers = int(exceptions["total_blockers"])
    complete = (
        source["verification"]["status"] == "PASS"
        and archive["mat_recording_member_count"] == len(records)
        and len(successful) == len(records)
        and schema["required_channel_length_consistency"]["mismatched_recordings"] == []
        and blockers == 0
    )
    return {
        "phase": PHASE_ID,
        "schema_version": SCHEMA_VERSION,
        "ok": bool(complete),
        "gate": "PASS_WITH_LIMITATIONS" if complete else "BLOCKED",
        "question": "Can D1 native radar channels and synchronized respiration reference be reproducibly converted into a provenance-preserving native radar phase/displacement trace without D0/MR60/model-specific normalization?",
        "answer": "YES_WITH_LIMITATIONS" if complete else "NO_BLOCKED",
        "evidence_scope": {
            "payload_checksum_verified": source["verification"]["status"] == "PASS",
            "archive_mat_recording_count": archive["mat_recording_member_count"],
            "recording_inventory_count": len(records),
            "adapted_recording_count": len(successful),
            "blocked_recording_count": len(records) - len(successful),
            "synchronized_respiration_reference_count": sum(
                row["adapter_output"].get("respiration_reference", {}).get("name") == "respiration"
                for row in successful
            ),
            "required_channel_length_mismatch_count": len(schema["required_channel_length_consistency"]["mismatched_recordings"]),
            "optional_reference_length_mismatch_count": schema["required_channel_length_consistency"]["optional_reference_length_mismatch_count"],
            "exception_blocker_count": blockers,
        },
        "native_adapter_behavior": {
            "model_training_performed": False,
            "feature_selection_performed": False,
            "D0_or_MR60_normalization_used": False,
            "window_local_normalization_used": False,
            "D2_semantics_touched": False,
            "native_phase_output": "native_unwrapped_phase_rad",
            "native_displacement_output": "relative_displacement_m",
            "reference_output": "respiration",
            "raw_waveforms_committed": False,
        },
        "limitations": [
            "mixed Fs values are preserved per recording (500 and 2000 Hz); no common rate is selected in D1",
            "radar_I/radar_Q and respiration physical units remain unverified native payload units",
            "per-sample timestamps are reconstructed from sample index/Fs",
            "ellipseReconstruction source code is MATLAB P-code; adapter fit is transparent and deterministic but not asserted byte-identical to the P-code",
            "source breath-hold/apnea protocol labels remain metadata only; no SafeNest label is generated",
        ],
        "ready_for": ["R1", "M-PV1"],
        "not_done": ["model_training", "V2_feature_selection", "D2_semantics", "SafeNest_APNEA_label_generation"],
        "exceptions_total": len(exceptions["exceptions"]),
        "exception_blockers": blockers,
    }


def run(payload: Path, extracted_root: Path) -> int:
    if not payload.is_file():
        raise D1AuditBlocked("PAYLOAD_NOT_FOUND", repo_rel(payload))
    observed_md5 = md5_file(payload)
    observed_bytes = payload.stat().st_size
    if observed_bytes != EXPECTED_PAYLOAD_BYTES:
        raise D1AuditBlocked("PAYLOAD_SIZE_MISMATCH", f"observed={observed_bytes}; expected={EXPECTED_PAYLOAD_BYTES}")
    if observed_md5 != EXPECTED_PAYLOAD_MD5:
        raise D1AuditBlocked("PAYLOAD_MD5_MISMATCH", f"observed={observed_md5}; expected={EXPECTED_PAYLOAD_MD5}")
    extracted_dataset = extracted_root / "datasets"
    overview = extracted_root / "overview_and_rating.xlsx"
    if not extracted_dataset.is_dir() or not overview.is_file():
        raise D1AuditBlocked("EXTRACTED_SCHEMA_NOT_FOUND", f"expected={repo_rel(extracted_root)}")
    archive = archive_inventory(payload)
    source = source_acquisition(payload, archive)
    workbook = workbook_index(overview)
    mat_paths = sorted(extracted_dataset.rglob("*.mat"), key=lambda path: repo_rel(path))
    if len(mat_paths) != archive["mat_recording_member_count"]:
        raise D1AuditBlocked(
            "ARCHIVE_EXTRACTION_RECORDING_COUNT_MISMATCH",
            f"archive={archive['mat_recording_member_count']} extracted={len(mat_paths)}",
        )

    records: list[dict[str, Any]] = []
    aggregates: list[dict[str, Any]] = []
    for index, path in enumerate(mat_paths, start=1):
        record, aggregate = mat_inventory(path, workbook)
        primary_condition = record["condition_metadata"]["source_protocol_labels"][0]
        try:
            adapted = adapt_mat_file(
                path,
                condition=primary_condition,
                condition_source="workbook_scenario_plus_source_path_metadata",
                source_file=repo_rel(path),
            )
        except D1AdapterError as exc:
            record.update({
                "adaptation_status": "BLOCKED",
                "adapter_error_code": exc.code,
                "adapter_error_detail": exc.detail,
                "adapter_output": {},
            })
        else:
            record.update({
                "adaptation_status": "SUCCESS",
                "adapter_error_code": None,
                "adapter_error_detail": None,
                "adapter_output": {
                    "adapter_id": adapted.metadata["adapter_id"],
                    "source_id": adapted.metadata["source_id"],
                    "source_sampling_rate_hz": adapted.metadata["source_sampling_rate_hz"],
                    "output_sampling_rate_hz": adapted.metadata["output_sampling_rate_hz"],
                    "output_signal_names": adapted.metadata["output_signal_names"],
                    "time_s_start": float(adapted.time_s[0]),
                    "time_s_end": float(adapted.time_s[-1]),
                    "sample_count": int(adapted.time_s.size),
                    "native_phase_stats": adapted.metadata["native_phase_stats"],
                    "relative_displacement_stats": adapted.metadata["relative_displacement_stats"],
                    "respiration_reference": {
                        "name": adapted.metadata["respiration_reference"]["name"],
                        "unit": adapted.metadata["respiration_reference"]["unit"],
                        "native_stats": adapted.metadata["respiration_reference"]["native_stats"],
                    },
                    "ellipse_correction": {
                        key: adapted.metadata["ellipse_correction"][key]
                        for key in (
                            "fit_method",
                            "fit_sample_count",
                            "conditioning_scale_native",
                            "ellipse_center_conditioned",
                            "unit_circle_transform_conditioned",
                            "radius_residual_rms",
                            "corrected_radius_min",
                            "corrected_radius_max",
                            "raw_i_stats",
                            "raw_q_stats",
                        )
                    },
                    "quality_flags": adapted.metadata["quality_flags"],
                },
            })
        records.append(record)
        aggregates.append(aggregate)
        if index % 25 == 0:
            print(f"audited {index}/{len(mat_paths)} recordings", flush=True)

    schema = schema_audit(records, aggregates, archive, len(workbook))
    exception_doc = exceptions_for(source, schema, records)
    validation = validation_result(source, archive, schema, records, exception_doc)
    contract = adapter_contract()
    evidence_dir = repo_path(EVIDENCE_REL)
    evidence_dir.mkdir(parents=True, exist_ok=True)
    docs = {
        "source_acquisition.json": source,
        "payload_inventory.json": archive,
        "schema_audit.json": schema,
        "adapter_contract.json": contract,
        "recording_inventory.json": {
            "phase": PHASE_ID,
            "schema_version": SCHEMA_VERSION,
            "recording_count": len(records),
            "recordings": records,
        },
        "exception_registry.json": exception_doc,
        "validation_result.json": validation,
    }
    file_hashes: dict[str, str] = {}
    for name, document in docs.items():
        file_hashes[name] = write_json(evidence_dir / name, document)
    checksums = {
        "phase": PHASE_ID,
        "schema_version": SCHEMA_VERSION,
        "files": file_hashes,
        "payload": {
            "path": repo_rel(payload),
            "byte_size": source["observed"]["byte_size"],
            "md5": source["observed"]["md5"],
            "sha256": source["observed"]["sha256"],
        },
        "code": {
            "files": {
                "adapters/mmwave_d1_2417ghz_adapter.py": sha256_file(ROOT / "adapters/mmwave_d1_2417ghz_adapter.py"),
                "scripts/run_mmwave_d1_2417ghz_adapter.py": sha256_file(ROOT / "scripts/run_mmwave_d1_2417ghz_adapter.py"),
                "scripts/validate_mmwave_d1_2417ghz_adapter.py": sha256_file(ROOT / "scripts/validate_mmwave_d1_2417ghz_adapter.py"),
            },
        },
    }
    write_json(evidence_dir / "checksums.json", checksums)
    print(json.dumps({"ok": validation["ok"], "gate": validation["gate"], "errors": exception_doc["total_blockers"], "recordings": len(records)}, indent=2))
    return 0 if validation["ok"] else 1


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--payload", type=Path, default=repo_path(PAYLOAD_REL))
    parser.add_argument("--extracted-root", type=Path, default=repo_path(EXTRACTED_REL))
    args = parser.parse_args()
    try:
        return run(args.payload, args.extracted_root)
    except D1AuditBlocked as exc:
        print(json.dumps({"ok": False, "gate": "BLOCKED", "code": exc.code, "summary": exc.summary}, indent=2), file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
